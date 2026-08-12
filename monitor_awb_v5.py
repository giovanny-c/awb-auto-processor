"""
Monitor de pasta para leitura automática de AWB (Air Waybill) via IA
=====================================================================

O QUE ESSE SCRIPT FAZ:
1. Fica observando uma pasta (onde o scanner salva os PDFs).
2. Quando um PDF novo aparece, espera ele terminar de ser gravado.
3. Detecta se o PDF é escaneado (sem texto) e, se necessário, aplica OCR.
4. Envia o PDF (+ texto OCR se houver) para a IA escolhida (Claude, Grok ou ChatGPT).
5. Grava uma nova linha na planilha xlsx com os dados extraídos.
6. Move o PDF processado para "processados" ou "erros".

COMO CONFIGURAR (antes de rodar):
1. Instale o Python 3.10+ e marque "Add Python to PATH".
2. Instale as bibliotecas:
       pip install watchdog anthropic openpyxl openai pymupdf pytesseract Pillow
3. Instale o Tesseract OCR no Windows:
   - Baixe em: https://github.com/UB-Mannheim/tesseract/wiki
   - Durante a instalação, anote o caminho (ex: C:\\Program Files\\Tesseract-OCR)
   - Adicione o idioma português se quiser (por padrão usa eng + por)
4. Crie as chaves de API:
   - ANTHROPIC_API_KEY  (Claude)
   - OPENAI_API_KEY     (ChatGPT)
   - XAI_API_KEY        (Grok)
5. Ajuste as configurações abaixo (pastas e TESSERACT_CMD se necessário).

COMO RODAR:
    no caminho do arquivo: python monitor_awb_v4.py

MODO DE TESTE:
No início do script você escolhe:
  - Teste  → gera dados aleatórios (não chama IA nem OCR)
  - Produção → usa a IA escolhida + OCR quando necessário
"""

# ==========================================================================
# IMPORTANDO DEPENDENCIAS
# ==========================================================================

import os
import sys
import json
import time
import random
import string
import shutil
import base64
import logging
from datetime import datetime
from pathlib import Path

# ==========================================================================
# CONFIGURAÇÕES DE AMBIENTE
# ==========================================================================

PASTA_ENTRADA      = r"C:\AWB\entrada"
PASTA_PROCESSADOS  = r"C:\AWB\processados"
PASTA_ERROS        = r"C:\AWB\erros"
PASTA_PLANILHAS    = r"C:\AWB\planilhas"
ARQUIVO_LOG        = r"C:\AWB\monitor_awb.log"

# Preenchido em tempo de execução pela escolha do usuário
TEST_MODE = False

# --- OCR ---
# True  → detecta PDFs escaneados e aplica Tesseract quando necessário
# False → envia o PDF direto para a IA (as IAs modernas já fazem OCR visual)
USAR_OCR = True

# Caminho do executável do Tesseract (deixe None para tentar detectar automaticamente)
# Exemplo Windows: r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESSERACT_CMD = None

# Idiomas do OCR (precisa ter os pacotes instalados no Tesseract)
# "por+eng" = português + inglês (recomendado para AWBs)
OCR_LANG = "por+eng"

# Limiar: se o PDF tiver menos que X caracteres de texto extraível → considera escaneado
MIN_CARACTERES_TEXTO = 80

# DPI usado na conversão de página para imagem (300 é bom equilíbrio qualidade/velocidade)
OCR_DPI = 300

# Modelos padrão
MODELOS = {
    "claude":  "claude-haiku-4-5-20251001",
    "grok":    "grok-4.6",
    "chatgpt": "gpt-4o",
}

ESPERA_ARQUIVO_ESTAVEL_SEGUNDOS = 3

# ==========================================================================
# CAMPOS E PROMPT
# ==========================================================================

CAMPOS_AWB = [
    "numero_awb",
    "shipper",
    "consignee",
    "aeroporto_origem",
    "aeroporto_destino",
    "gross_weight",
    "chargeable_weight",
    "volumes_quantity",
    "descricao_mercadoria",
    "valor_declarado",
    "valor_frete",
]

CABECALHO_PLANILHA = [
    "Data/Hora Processamento",
    "Arquivo Origem",
    "Número AWB",
    "Shipper (Expedidor)",
    "Consignatário",
    "Aeroporto Origem",
    "Aeroporto Destino",
    "Peso Bruto",
    "Peso Taxado",
    "Qtd. Volumes",
    "Descrição Mercadoria",
    "Valor Declarado",
    "Valor Frete",
]

PROMPT_EXTRACAO = """Você está vendo uma AWB (Air Waybill / Conhecimento Aéreo) digitalizada.

Extraia EXATAMENTE os seguintes campos do documento e responda SOMENTE com um
objeto JSON válido, sem nenhum texto antes ou depois, sem markdown, sem ```.

Campos a extrair:
- numero_awb: número da AWB (geralmente no formato XXX-XXXXXXXX se for uma MAWB e de 5 a 15 caracteres alfa-numéricos se for uma HAWB)
- shipper: nome do expedidor/remetente
- consignee: nome do consignatário/destinatário
- aeroporto_origem: código do aeroporto de origem 
- aeroporto_destino: código do aeroporto de destino
- gross_weight: peso bruto da carga (com unidade, ex: "320 kg")
- chargeable_weight: campo chargeable weight (com unidade)
- volumes_quantity: quantidade de volumes/peças
- descricao_mercadoria: descrição resumida da mercadoria
- valor_declarado: valor declarado para transporte
- valor_frete: valor do frete (campos de other charges)

Se algum campo não estiver visível ou legível no documento, use o valor null
para esse campo. Não invente informações. Responda apenas com o JSON.

Formato esperado:
{"numero_awb": "...", "shipper": "...", "consignee": "...", "gross_weight": "...", 
 "aeroporto_origem": "...", "aeroporto_destino": "...",
 "chargeable_weight": "...", "volumes_quantity": "...",
 "descricao_mercadoria": "...", "valor_declarado": "...", "valor_frete": "..."}
"""

# ==========================================================================
# LOGGING E PASTAS
# ==========================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(ARQUIVO_LOG, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("monitor_awb")


def garantir_pastas():
    for pasta in (PASTA_ENTRADA, PASTA_PROCESSADOS, PASTA_ERROS, PASTA_PLANILHAS):
        Path(pasta).mkdir(parents=True, exist_ok=True)


def caminho_planilha_do_dia() -> Path:
    nome_arquivo = f"awb_dados_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return Path(PASTA_PLANILHAS) / nome_arquivo


def esperar_arquivo_estavel(caminho: Path, espera: int = ESPERA_ARQUIVO_ESTAVEL_SEGUNDOS):
    tamanho_anterior = -1
    tentativas_sem_mudanca = 0
    while tentativas_sem_mudanca < 2:
        try:
            tamanho_atual = caminho.stat().st_size
        except FileNotFoundError:
            time.sleep(1)
            continue
        if tamanho_atual == tamanho_anterior and tamanho_atual > 0:
            tentativas_sem_mudanca += 1
        else:
            tentativas_sem_mudanca = 0
        tamanho_anterior = tamanho_atual
        time.sleep(espera)


# ==========================================================================
# OCR — DETECÇÃO + EXTRAÇÃO
# ==========================================================================

def pdf_tem_texto_suficiente(caminho_pdf: Path) -> bool:
    """
    Verifica se o PDF já possui camada de texto suficiente.
    Retorna True se tiver texto legível, False se parecer escaneado.
    """
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(caminho_pdf)
        total_chars = 0
        for page in doc:
            total_chars += len(page.get_text("text").strip())
        doc.close()
        log.info(f"Texto extraível no PDF: {total_chars} caracteres")
        return total_chars >= MIN_CARACTERES_TEXTO
    except Exception as e:
        log.warning(f"Não foi possível analisar o texto do PDF: {e}")
        return False  # em caso de dúvida, tenta OCR


def extrair_texto_ocr(caminho_pdf: Path) -> str:
    """
    Converte cada página do PDF em imagem e roda Tesseract OCR.
    Retorna o texto concatenado de todas as páginas.
    """
    try:
        import fitz
        import pytesseract
        from PIL import Image
        import io

        # Configura o caminho do Tesseract se informado
        if TESSERACT_CMD:
            pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

        doc = fitz.open(caminho_pdf)
        textos = []

        for i, page in enumerate(doc):
            # Renderiza a página em alta resolução
            mat = fitz.Matrix(OCR_DPI / 72, OCR_DPI / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_bytes))

            texto_pagina = pytesseract.image_to_string(img, lang=OCR_LANG)
            textos.append(f"--- Página {i + 1} ---\n{texto_pagina.strip()}")
            log.info(f"OCR página {i + 1}/{len(doc)} concluído")

        doc.close()
        texto_completo = "\n\n".join(textos)
        log.info(f"OCR finalizado. Total de caracteres extraídos: {len(texto_completo)}")
        return texto_completo

    except Exception as e:
        log.error(f"Falha no OCR: {e}")
        raise RuntimeError(
            f"Erro ao executar OCR. Verifique se o Tesseract está instalado "
            f"e se o caminho TESSERACT_CMD está correto. Detalhe: {e}"
        ) from e


def preparar_conteudo_pdf(caminho_pdf: Path) -> tuple[str, str | None]:
    """
    Decide se precisa de OCR e retorna:
    - pdf_base64
    - texto_ocr (ou None se não precisar / não usar OCR)
    """
    with open(caminho_pdf, "rb") as f:
        pdf_base64 = base64.standard_b64encode(f.read()).decode("utf-8")

    texto_ocr = None

    if USAR_OCR and not TEST_MODE:
        if pdf_tem_texto_suficiente(caminho_pdf):
            log.info("PDF já possui texto legível → OCR não necessário")
        else:
            log.info("PDF parece escaneado (pouco ou nenhum texto) → aplicando OCR...")
            texto_ocr = extrair_texto_ocr(caminho_pdf)

    return pdf_base64, texto_ocr


# ==========================================================================
# DADOS DE TESTE
# ==========================================================================

AEROPORTOS_TESTE = ["GRU", "GIG", "VCP", "CGH", "REC", "SSA", "CNF", "POA", "MIA", "JFK", "MAD", "LIS"]
MERCADORIAS_TESTE = [
    "Peças automotivas",
    "Equipamentos eletrônicos",
    "Produtos farmacêuticos",
    "Têxteis diversos",
    "Máquinas industriais",
    "Amostras sem valor comercial",
]


def gerar_dados_teste_aleatorios(caminho_pdf: Path) -> dict:
    origem, destino = random.sample(AEROPORTOS_TESTE, 2)
    gross_weight = round(random.uniform(10, 5000), 1)
    chargeable_weight = round(gross_weight * random.uniform(1.0, 1.2), 1)
    numero_awb = f"{random.randint(100, 999)}-{''.join(random.choices(string.digits, k=8))}"

    dados = {
        "numero_awb": numero_awb,
        "shipper": f"{random.choice(['Alfa', 'Beta', 'Global', 'Nova', 'Sul', 'Norte'])} Exportadora Ltda",
        "consignee": f"{random.choice(['Prime', 'Rio', 'Vale', 'Atlas', 'Delta', 'Bravo'])} Importadora Ltda",
        "aeroporto_origem": origem,
        "aeroporto_destino": destino,
        "gross_weight": f"{gross_weight} kg",
        "chargeable_weight": f"{chargeable_weight} kg",
        "volumes_quantity": random.randint(1, 50),
        "descricao_mercadoria": random.choice(MERCADORIAS_TESTE),
        "valor_declarado": round(random.uniform(500, 50000), 2),
        "valor_frete": round(random.uniform(100, 5000), 2),
    }
    log.info(f"[TESTE] Dados aleatórios gerados para {caminho_pdf.name}: {dados['numero_awb']}")
    return dados


# ==========================================================================
# EXTRAÇÃO POR PROVEDOR DE IA
# ==========================================================================

def _limpar_resposta_json(texto: str) -> str:
    texto = texto.strip()
    if texto.startswith("```"):
        texto = texto.strip("`")
        if texto.lower().startswith("json"):
            texto = texto[4:].strip()
    return texto


def _montar_prompt_com_ocr(texto_ocr: str | None) -> str:
    """Adiciona o texto OCR ao prompt quando disponível."""
    if not texto_ocr:
        return PROMPT_EXTRACAO

    return (
        PROMPT_EXTRACAO
        + "\n\n--- TEXTO EXTRAÍDO VIA OCR (use como apoio, priorize o que está visível no documento) ---\n"
        + texto_ocr
        + "\n--- FIM DO TEXTO OCR ---\n"
    )


def extrair_com_claude(caminho_pdf: Path, pdf_base64: str, texto_ocr: str | None) -> dict:
    from anthropic import Anthropic

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY não encontrada.")

    client = Anthropic(api_key=api_key)
    prompt = _montar_prompt_com_ocr(texto_ocr)

    resposta = client.messages.create(
        model=MODELOS["claude"],
        max_tokens=1024,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_base64,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }
        ],
    )

    texto_resposta = "".join(
        bloco.text for bloco in resposta.content if bloco.type == "text"
    )
    texto_resposta = _limpar_resposta_json(texto_resposta)

    try:
        return json.loads(texto_resposta)
    except json.JSONDecodeError as e:
        raise ValueError(f"Claude não retornou JSON válido. Resposta: {texto_resposta[:500]}") from e


def extrair_com_openai_compativel(
    caminho_pdf: Path,
    pdf_base64: str,
    texto_ocr: str | None,
    provedor: str,
) -> dict:
    from openai import OpenAI

    if provedor == "chatgpt":
        api_key = os.environ.get("OPENAI_API_KEY")
        if not api_key:
            raise RuntimeError("OPENAI_API_KEY não encontrada.")
        client = OpenAI(api_key=api_key)
        modelo = MODELOS["chatgpt"]
    elif provedor == "grok":
        api_key = os.environ.get("XAI_API_KEY")
        if not api_key:
            raise RuntimeError("XAI_API_KEY não encontrada.")
        client = OpenAI(api_key=api_key, base_url="https://api.x.ai/v1")
        modelo = MODELOS["grok"]
    else:
        raise ValueError(f"Provedor desconhecido: {provedor}")

    prompt = _montar_prompt_com_ocr(texto_ocr)

    content = [
        {
            "type": "file",
            "file": {
                "filename": caminho_pdf.name,
                "file_data": f"data:application/pdf;base64,{pdf_base64}",
            },
        },
        {
            "type": "text",
            "text": prompt,
        },
    ]

    resposta = client.chat.completions.create(
        model=modelo,
        max_tokens=1024,
        messages=[{"role": "user", "content": content}],
    )

    texto_resposta = resposta.choices[0].message.content or ""
    texto_resposta = _limpar_resposta_json(texto_resposta)

    try:
        return json.loads(texto_resposta)
    except json.JSONDecodeError as e:
        raise ValueError(
            f"{provedor.upper()} não retornou JSON válido. Resposta: {texto_resposta[:500]}"
        ) from e


def extrair_dados_awb(caminho_pdf: Path, provedor: str) -> dict:
    """Função principal: prepara o PDF (com OCR se necessário) e chama a IA."""
    pdf_base64, texto_ocr = preparar_conteudo_pdf(caminho_pdf)

    provedor = provedor.lower().strip()

    if provedor == "claude":
        return extrair_com_claude(caminho_pdf, pdf_base64, texto_ocr)
    elif provedor in ("grok", "chatgpt"):
        return extrair_com_openai_compativel(caminho_pdf, pdf_base64, texto_ocr, provedor)
    else:
        raise ValueError(f"Provedor inválido: '{provedor}'. Use 'claude', 'grok' ou 'chatgpt'.")


# ==========================================================================
# PLANILHA E PROCESSAMENTO
# ==========================================================================

def gravar_na_planilha(caminho_pdf: Path, dados: dict):
    from openpyxl import Workbook, load_workbook

    caminho_planilha = caminho_planilha_do_dia()

    if caminho_planilha.exists():
        wb = load_workbook(caminho_planilha)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AWBs"
        ws.append(CABECALHO_PLANILHA)
        for col_idx, titulo in enumerate(CABECALHO_PLANILHA, start=1):
            ws.cell(row=1, column=col_idx).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)

    linha = [
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        caminho_pdf.name,
    ] + [dados.get(campo) for campo in CAMPOS_AWB]

    ws.append(linha)
    wb.save(caminho_planilha)


def processar_arquivo(caminho_pdf: Path, provedor: str):
    log.info(f"Novo arquivo detectado: {caminho_pdf.name}")
    try:
        esperar_arquivo_estavel(caminho_pdf)

        if TEST_MODE:
            dados = gerar_dados_teste_aleatorios(caminho_pdf)
        else:
            dados = extrair_dados_awb(caminho_pdf, provedor)

        gravar_na_planilha(caminho_pdf, dados)

        destino = Path(PASTA_PROCESSADOS) / caminho_pdf.name
        shutil.move(str(caminho_pdf), str(destino))
        log.info(f"Processado com sucesso: {caminho_pdf.name} → {dados.get('numero_awb')}")

    except Exception as e:
        log.error(f"Erro ao processar {caminho_pdf.name}: {e}")
        try:
            destino_erro = Path(PASTA_ERROS) / caminho_pdf.name
            shutil.move(str(caminho_pdf), str(destino_erro))
        except Exception as e2:
            log.error(f"Não foi possível mover o arquivo com erro: {e2}")


def rodar_varredura_inicial(provedor: str):
    for arquivo in Path(PASTA_ENTRADA).glob("*.pdf"):
        processar_arquivo(arquivo, provedor)


def escolher_modo_teste() -> bool:
    """Pergunta ao usuário se deseja rodar em modo de teste ou produção."""
    print("\n" + "=" * 50)
    print("  Escolha o modo de execução:")
    print("=" * 50)
    print("  1. Teste      → gera dados aleatórios (não chama IA nem OCR)")
    print("  2. Produção   → usa IA + OCR quando necessário")
    print("=" * 50)

    while True:
        escolha = input("Digite 1 ou 2 (ou: teste / producao / produção): ").strip().lower()
        mapa = {
            "1": True,
            "2": False,
            "teste": True,
            "test": True,
            "t": True,
            "producao": False,
            "produção": False,
            "prod": False,
            "p": False,
        }
        if escolha in mapa:
            modo_teste = mapa[escolha]
            if modo_teste:
                print("\n→ Modo selecionado: TESTE (dados aleatórios)\n")
            else:
                print("\n→ Modo selecionado: PRODUÇÃO (IA real)\n")
            return modo_teste
        print("Opção inválida. Tente novamente.")


def escolher_provedor() -> str:
    print("\n" + "=" * 50)
    print("  Escolha a IA que será usada para ler as AWBs:")
    print("=" * 50)
    print("  1. Claude   (Anthropic)")
    print("  2. Grok     (xAI)")
    print("  3. ChatGPT  (OpenAI)")
    print("=" * 50)

    while True:
        escolha = input("Digite 1, 2 ou 3 (ou o nome: claude / grok / chatgpt): ").strip().lower()
        mapa = {
            "1": "claude", "2": "grok", "3": "chatgpt",
            "claude": "claude", "grok": "grok", "chatgpt": "chatgpt",
            "openai": "chatgpt", "xai": "grok", "anthropic": "claude",
        }
        if escolha in mapa:
            provedor = mapa[escolha]
            print(f"\n→ IA selecionada: {provedor.upper()}\n")
            return provedor
        print("Opção inválida. Tente novamente.")


def main():
    global TEST_MODE

    garantir_pastas()

    TEST_MODE = escolher_modo_teste()

    if TEST_MODE:
        provedor = "teste"
        log.warning("Modo TESTE ativo → gerando dados ALEATÓRIOS (IA e OCR desativados).")
    else:
        provedor = escolher_provedor()
        log.info(f"Provedor de IA selecionado: {provedor}")
        if USAR_OCR:
            log.info("OCR ativado: PDFs escaneados serão processados com Tesseract.")
        else:
            log.info("OCR desativado: o PDF será enviado direto para a IA.")

    log.info("=" * 60)
    log.info("Monitor de AWB iniciado")
    log.info(f"Modo: {'TESTE' if TEST_MODE else 'PRODUÇÃO'}")
    log.info(f"Observando pasta: {PASTA_ENTRADA}")
    log.info(f"Pasta de planilhas: {PASTA_PLANILHAS}")
    if not TEST_MODE:
        log.info(f"Modelo: {MODELOS.get(provedor, 'N/A')}")
    log.info("=" * 60)

    rodar_varredura_inicial(provedor)

    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler

    class ManipuladorPDF(FileSystemEventHandler):
        def on_created(self, event):
            if event.is_directory:
                return
            if event.src_path.lower().endswith(".pdf"):
                processar_arquivo(Path(event.src_path), provedor)

    observador = Observer()
    observador.schedule(ManipuladorPDF(), PASTA_ENTRADA, recursive=False)
    observador.start()

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observador.stop()
        log.info("Monitor encerrado pelo usuário.")
    observador.join()


if __name__ == "__main__":
    main()
